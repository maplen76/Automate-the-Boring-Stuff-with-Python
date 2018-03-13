
#! python3
# multiplicationTable.py - Create NxN multiplication

import sys, os, openpyxl
from openpyxl.styles import Font

os.chdir('D:\\PythonLab')

N = int(sys.argv[1]) # sys.argv[0] is the name of the script

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, N + 1):
    boldFont = Font(bold = True)
    sheet.cell(row = 1, column = i + 1).value = i    
    sheet.cell(row = 1, column = i + 1).font = boldFont
    
    for j in range(1, N + 1):
        sheet.cell(row = j+1, column = 1).value = j
        sheet.cell(row = j+1, column = 1).font = boldFont

        sheet.cell(row = i + 1, column = j + 1).value = i*j   
    
wb.save('test.xlsx')
